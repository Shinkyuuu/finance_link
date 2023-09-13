from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException

from openpyxl import load_workbook
import os
from time import sleep
import json
from dotenv import load_dotenv

load_dotenv()

class amexDriver:
    def __init__(self):
        self.prepareDownloadDir()

        self.options = Options()
        self.options.add_argument('--ignore-certificate-errors-spki-list')
        self.options.add_argument('--ignore-ssl-errors')
        self.options.add_argument('--headless=new')

        self.options.add_experimental_option('prefs',  {
            "download.default_directory": self.download_dir,
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "plugins.always_open_pdf_externally": True
            }
        )
        
        self.driver = webdriver.Chrome(service = Service(ChromeDriverManager().install()), options = self.options)
        self.driver.get('https://www.americanexpress.com/en-us/account/login?inav=iNavLnkLog')

        self.wait = WebDriverWait(self.driver, 60)

        return
    
    def waitClick(self, newBy, newValue):

        ignored_exceptions=(NoSuchElementException,StaleElementReferenceException)
        button = WebDriverWait(self.driver, 10, ignored_exceptions=ignored_exceptions) \
            .until(EC.presence_of_element_located((newBy, newValue)))

        # self.wait.until(EC.element_to_be_clickable((newBy, newValue)))
        # button = self.driver.find_element(by = newBy, value = newValue)
        self.driver.execute_script("arguments[0].click();", button)
        
        return
    
    def prepareDownloadDir(self):
        self.download_dir = os.path.expanduser('~\Documents') + "\\finlink\\amex\\"

        if not os.path.exists(self.download_dir):
            os.makedirs(self.download_dir)

        print(f"download path set to {self.download_dir}")

        for activity in os.listdir(self.download_dir):
            # if not activity.endswith(".xlsx"):
            #     continue
            os.remove(os.path.join(self.download_dir, activity))

        return

    def login(self):
        usernameInput = self.driver.find_element(by = By.ID, value="eliloUserID")
        passwordInput = self.driver.find_element(by = By.ID, value="eliloPassword")
        usernameInput.send_keys(os.getenv("AMEX_USERNAME")) # HIDE PERSONAL INFORMATION
        passwordInput.send_keys(os.getenv("AMEX_PASSWORD")) # HIDE PERSONAL INFORMATION
        
        self.waitClick(By.ID, "loginSubmit")

        return

    def getActivity(self):
        self.waitClick(By.XPATH,  "//a[@title='Statements & Activity']")

        self.waitClick(By.CLASS_NAME,  "account-switcher-toggler")

        # Acquire the lsit of credit accounts on the AMEX profile
        # self.waitClick(By.ID,  "switcher_product_rows")

        accountToggleList = self.driver.find_element(by = By.ID, value="switcher_product_rows")
        accountList = accountToggleList.find_elements(By.CSS_SELECTOR, 'li')

        if not accountList:
            print("ERROR") # Handle Error
            return

        self.numOfAccounts = len(accountList)

        for i in range(self.numOfAccounts):
            if i != 0:
                self.waitClick(By.CLASS_NAME,  "account-switcher-toggler")

                # Special case
                accountToggleList = self.driver.find_element(by = By.ID, value="switcher_product_rows")
                accountToggleList.find_elements(By.CSS_SELECTOR, 'li')[i].click()
            print(f"Getting activity from account {i + 1}")

            # Click on the 'download activity' button
            self.waitClick(By.XPATH,  "//i[@class='btn btn-icon transparent icon dls-icon-download icon-hover margin-2-l-md-up dls-bright-blue ']")
            self.waitClick(By.XPATH,  "//label[@for='axp-activity-download-body-selection-options-type_excel']")

            # Select the 'excel' and 'select all' toggles

            allOption = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//label[@for='axp-activity-download-body-checkbox-options-includeAll']")))

            if not allOption.get_attribute('checked'):
                allOption.click()
            
            # Click on the 'download' button
            self.waitClick(By.XPATH,  "//a[@label='Download']")

            # sleep(0.5)
        return

    # Open and Iterate through each activity excel sheet
    def getActivitySheets(self):
        for activity in os.listdir(self.download_dir):
            if not activity.endswith(".xlsx"):
                continue
            # os.remove(os.path.join(self.download_dir, activity))
            print()
            print(f"openning {self.download_dir + activity}")
            workbook = load_workbook(filename = (self.download_dir + activity))
            self.parseActivity(workbook)

    #Convert an activity spreadsheet into json
    def parseActivity(self, workbook):
        activityDict = {
            "transHistorySheet" : [],
            "balanceInfoSheet" : {}
        }

        transHistorySheet = workbook["Transaction Details"]
        balanceInfoSheet = workbook["Transaction Summary"]

        for date, description, amount in transHistorySheet.iter_rows(min_row = 8, min_col = 1, max_col = 3, max_row = transHistorySheet.max_row):
            if date.value is None: break

            activityDict["transHistorySheet"].append({
                "date":             date.value,
                "description":  description.value,
                "amount":        amount.value
            })

        activityDict["balanceInfoSheet"] = {
            "prevBalance":                  balanceInfoSheet['B9'].value,
            "paymentsAndCredits":    balanceInfoSheet['B10'].value,
            "charges":                         balanceInfoSheet['B11'].value,
            "totalBalance":                  balanceInfoSheet['B12'].value
        }            

        print(json.dumps(activityDict))
        return 

myAmex = amexDriver()
myAmex.login()
myAmex.getActivity()
myAmex.getActivitySheets()

sleep(100)